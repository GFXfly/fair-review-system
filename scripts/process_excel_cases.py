#!/usr/bin/env python3
"""
处理 Excel 案例数据，进行脱敏处理并生成 TypeScript 导入脚本
"""

import openpyxl
import re
import json
import os

# 脱敏处理函数
def anonymize_text(text):
    """对文本进行脱敏处理，保留省级，模糊化市/区/县级"""
    if not text:
        return text
    
    result = text
    
    # 地级市脱敏（保留省份）
    cities = [
        '杭州市', '宁波市', '温州市', '嘉兴市', '湖州市', '绍兴市', '金华市', 
        '衢州市', '舟山市', '台州市', '丽水市',  # 浙江省地级市
        '南京市', '苏州市', '无锡市', '常州市', '镇江市', '南通市', '扬州市',
        '盐城市', '徐州市', '淮安市', '连云港市', '泰州市', '宿迁市',  # 江苏省
        '广州市', '深圳市', '珠海市', '佛山市', '东莞市', '中山市', '惠州市',
        '汕头市', '湛江市', '茂名市', '肇庆市', '梅州市', '汕尾市', '河源市',
        '阳江市', '清远市', '潮州市', '揭阳市', '云浮市', '韶关市', '江门市',  # 广东省
        '成都市', '绵阳市', '德阳市', '攀枝花市', '泸州市', '广元市', '遂宁市',
        '内江市', '乐山市', '南充市', '眉山市', '宜宾市', '广安市', '达州市',
        '雅安市', '巴中市', '资阳市', '自贡市',  # 四川省
        '济南市', '青岛市', '烟台市', '潍坊市', '威海市', '临沂市', '淄博市',
        '济宁市', '泰安市', '聊城市', '德州市', '滨州市', '东营市', '菏泽市',
        '枣庄市', '日照市',  # 山东省
        '福州市', '厦门市', '泉州市', '漳州市', '莆田市', '龙岩市', '三明市',
        '南平市', '宁德市',  # 福建省
        '武汉市', '宜昌市', '襄阳市', '荆州市', '黄石市', '十堰市', '荆门市',
        '鄂州市', '孝感市', '黄冈市', '咸宁市', '随州市',  # 湖北省
        '长沙市', '株洲市', '湘潭市', '衡阳市', '邵阳市', '岳阳市', '常德市',
        '张家界市', '益阳市', '娄底市', '郴州市', '永州市', '怀化市',  # 湖南省
        '太原市', '大同市', '阳泉市', '长治市', '晋城市', '朔州市', '晋中市',
        '运城市', '忻州市', '临汾市', '吕梁市',  # 山西省
        '西安市', '咸阳市', '宝鸡市', '渭南市', '铜川市', '延安市', '榆林市',
        '汉中市', '安康市', '商洛市',  # 陕西省
        '合肥市', '芜湖市', '蚌埠市', '淮南市', '马鞍山市', '淮北市', '铜陵市',
        '安庆市', '黄山市', '滁州市', '阜阳市', '宿州市', '六安市', '亳州市',
        '池州市', '宣城市',  # 安徽省
        '南昌市', '九江市', '景德镇市', '萍乡市', '新余市', '鹰潭市', '赣州市',
        '吉安市', '宜春市', '抚州市', '上饶市',  # 江西省
        '郑州市', '开封市', '洛阳市', '平顶山市', '安阳市', '鹤壁市', '新乡市',
        '焦作市', '濮阳市', '许昌市', '漯河市', '三门峡市', '南阳市', '商丘市',
        '信阳市', '周口市', '驻马店市',  # 河南省
        '石家庄市', '唐山市', '秦皇岛市', '邯郸市', '邢台市', '保定市', '张家口市',
        '承德市', '沧州市', '廊坊市', '衡水市',  # 河北省
        '哈尔滨市', '齐齐哈尔市', '牡丹江市', '佳木斯市', '大庆市', '鸡西市',
        '双鸭山市', '伊春市', '七台河市', '鹤岗市', '黑河市', '绥化市',  # 黑龙江省
        '长春市', '吉林市', '四平市', '辽源市', '通化市', '白山市', '松原市',
        '白城市',  # 吉林省
        '沈阳市', '大连市', '鞍山市', '抚顺市', '本溪市', '丹东市', '锦州市',
        '营口市', '阜新市', '辽阳市', '盘锦市', '铁岭市', '朝阳市', '葫芦岛市',  # 辽宁省
        '呼和浩特市', '包头市', '乌海市', '赤峰市', '通辽市', '鄂尔多斯市',
        '呼伦贝尔市', '巴彦淖尔市', '乌兰察布市',  # 内蒙古
        '南宁市', '柳州市', '桂林市', '梧州市', '北海市', '防城港市', '钦州市',
        '贵港市', '玉林市', '百色市', '贺州市', '河池市', '来宾市', '崇左市',  # 广西
        '海口市', '三亚市', '三沙市', '儋州市',  # 海南
        '贵阳市', '六盘水市', '遵义市', '安顺市', '毕节市', '铜仁市',  # 贵州
        '昆明市', '曲靖市', '玉溪市', '保山市', '昭通市', '丽江市', '普洱市',
        '临沧市',  # 云南
        '拉萨市', '日喀则市', '昌都市', '林芝市', '山南市', '那曲市',  # 西藏
        '兰州市', '嘉峪关市', '金昌市', '白银市', '天水市', '武威市', '张掖市',
        '平凉市', '酒泉市', '庆阳市', '定西市', '陇南市',  # 甘肃
        '西宁市', '海东市',  # 青海
        '银川市', '石嘴山市', '吴忠市', '固原市', '中卫市',  # 宁夏
        '乌鲁木齐市', '克拉玛依市', '吐鲁番市', '哈密市',  # 新疆
        '北京市', '天津市', '上海市', '重庆市',  # 直辖市
        '临海市', '义乌市', '东阳市', '永康市', '江山市', '龙泉市',  # 县级市
    ]
    
    for city in cities:
        if city in result:
            result = result.replace(city, '某市')
    
    # 区县脱敏
    # 匹配"XX区"、"XX县"、"XX市"（县级市）
    result = re.sub(r'([临平|拱墅|西湖|上城|下城|江干|滨江|萧山|余杭|富阳|桐庐|建德|淳安|鹿城|龙湾|瓯海|洞头|瑞安|乐清|永嘉|平阳|苍南|文成|泰顺|南湖|秀洲|嘉善|海盐|海宁|平湖|桐乡|吴兴|南浔|德清|长兴|安吉|越城|柯桥|上虞|诸暨|嵊州|新昌|婺城|金东|兰溪|东阳|永康|武义|浦江|磐安|柯城|衢江|江山|龙游|常山|开化|定海|普陀|岱山|嵊泗|椒江|黄岩|路桥|三门|天台|仙居|温岭|临海|玉环|莲都|龙泉|青田|缙云|遂昌|松阳|云和|庆元|景宁]+)(区|县)', r'某\2', result)
    
    # 更通用的区县匹配
    result = re.sub(r'([一-龥]{2,4})(区|县)(?=[^\u4e00-\u9fa5]|$|[。，、；：""''（）])', r'某\2', result)
    
    # 具体街道/乡镇/村
    result = re.sub(r'([一-龥]{2,6})(街道|乡|镇|村)(?=[^\u4e00-\u9fa5]|$|[。，、；：""''（）])', r'某\2', result)
    
    # 清理多余的"某某"
    result = result.replace('某某', '某')
    
    return result

def extract_violation_type(text):
    """从违规类型字段提取简化的违规类型"""
    if not text:
        return "其他违规"
    
    type_mapping = {
        '市场准入': '设置不合理的市场准入条件',
        '准入和退出': '设置不合理的市场准入条件',
        '商品和要素自由流动': '妨碍商品和要素自由流动',
        '影响生产经营成本': '给予特定经营者差异化财政奖励',
        '影响生产经营行为': '影响生产经营行为',
    }
    
    for key, value in type_mapping.items():
        if key in text:
            return value
    
    return "其他违规"

def process_file1(filepath):
    """处理文件1: 全国公平竞争审查督查发现的问题政策措施"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    cases = []
    # 从第3行开始读取数据（第1行标题，第2行表头）
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:  # 跳过空行
            continue
        
        seq = row[0]  # 序号
        doc_name = row[1]  # 文件名称
        org = row[2]  # 发文单位
        content = row[3]  # 文件有关内容
        violation_type_raw = row[4]  # 违规类型
        violation_clause = row[5]  # 违反的标准
        review_opinion = row[6]  # 审查意见
        
        if not doc_name or not content:
            continue
        
        # 脱敏处理
        doc_name_anon = anonymize_text(doc_name)
        org_anon = anonymize_text(org.replace('\n', '') if org else '')
        content_anon = anonymize_text(content)
        review_opinion_anon = anonymize_text(review_opinion) if review_opinion else ''
        
        case = {
            'title': f"{org_anon}发布的{doc_name_anon}涉嫌违反公平竞争审查规定",
            'content': f"{review_opinion_anon}\n\n【违规条款原文】\n{content_anon}",
            'violationType': extract_violation_type(violation_type_raw),
            'result': '责令整改',
            'violationClause': violation_clause.replace('\n', ' ') if violation_clause else '',
            'documentName': doc_name_anon,
            'documentOrg': org_anon,
            'province': '浙江省',
            'violationDetail': content_anon[:200] if content_anon else '',
        }
        cases.append(case)
    
    return cases

def process_file2(filepath):
    """处理文件2: 浙江省公平竞争审查第三方评估发现的妨碍统一市场和公平竞争问题"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    cases = []
    # 从第3行开始读取数据
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        
        seq = row[0]  # 序号
        doc_name = row[1]  # 文件名称
        org = row[2]  # 发文单位
        pub_date = row[3]  # 发文时间
        violation_content = row[4]  # 涉嫌违规内容
        violation_clause = row[5]  # 涉嫌违反的规定
        
        if not doc_name or not violation_content:
            continue
        
        # 脱敏处理
        doc_name_anon = anonymize_text(doc_name)
        org_anon = anonymize_text(org) if org else ''
        violation_content_anon = anonymize_text(violation_content)
        violation_clause_str = violation_clause if violation_clause else ''
        
        # 提取违规类型
        violation_type = '其他违规'
        if '市场准入' in violation_clause_str:
            violation_type = '设置不合理的市场准入条件'
        elif '商品和要素自由流动' in violation_clause_str:
            violation_type = '妨碍商品和要素自由流动'
        elif '招标投标' in violation_clause_str or '投标' in violation_content:
            violation_type = '排斥或限制外地经营者参加本地招标投标'
        elif '注册' in violation_content or '分支机构' in violation_content:
            violation_type = '变相强制外地经营者在本地设立分支机构'
        
        case = {
            'title': f"{org_anon}发布的{doc_name_anon}涉嫌违反公平竞争审查规定",
            'content': f"{violation_content_anon}\n\n【涉嫌违反的规定】\n{violation_clause_str[:500] if violation_clause_str else ''}",
            'violationType': violation_type,
            'result': '责令整改',
            'violationClause': violation_clause_str[:200] if violation_clause_str else '',
            'documentName': doc_name_anon,
            'documentOrg': org_anon,
            'province': '浙江省',
            'violationDetail': violation_content_anon[:200] if violation_content_anon else '',
            'publishDate': str(pub_date) if pub_date else None,
        }
        cases.append(case)
    
    return cases

def generate_ts_script(cases, report_title, report_dept, report_date, output_file):
    """生成 TypeScript 导入脚本"""
    
    # 转义字符串中的特殊字符
    def escape_ts_string(s):
        if not s:
            return ''
        return s.replace('\\', '\\\\').replace("'", "\\'").replace('\n', '\\n').replace('\r', '')
    
    ts_content = f'''import {{ PrismaClient }} from '@prisma/client';

const prisma = new PrismaClient();

async function main() {{
    // 1. 创建通报记录
    let report = await prisma.report.findFirst({{
        where: {{ title: '{escape_ts_string(report_title)}' }}
    }});

    if (!report) {{
        report = await prisma.report.create({{
            data: {{
                title: '{escape_ts_string(report_title)}',
                department: '{escape_ts_string(report_dept)}',
                publishDate: '{report_date}',
                province: '浙江省'
            }}
        }});
        console.log(`创建通报记录，ID: ${{report.id}}`);
    }} else {{
        console.log(`找到已有通报记录，ID: ${{report.id}}`);
    }}

    // 2. 添加案例
    const cases = [
'''
    
    for case in cases:
        ts_content += f'''        {{
            title: '{escape_ts_string(case["title"][:150])}',
            content: '{escape_ts_string(case["content"])}',
            violationType: '{escape_ts_string(case["violationType"])}',
            result: '{escape_ts_string(case["result"])}',
            violationClause: '{escape_ts_string(case.get("violationClause", "")[:200])}',
            documentName: '{escape_ts_string(case["documentName"])}',
            documentOrg: '{escape_ts_string(case["documentOrg"])}',
            province: '{escape_ts_string(case["province"])}',
            violationDetail: '{escape_ts_string(case.get("violationDetail", ""))}',
            reportId: report.id,
            publishDate: '{report_date}'
        }},
'''
    
    ts_content += '''    ];

    // 批量创建案例
    let addedCount = 0;
    for (const c of cases) {
        const existing = await prisma.case.findFirst({
            where: { title: c.title }
        });

        if (!existing) {
            await prisma.case.create({ data: c });
            addedCount++;
            console.log(`✓ 添加案例 ${addedCount}: ${c.title.substring(0, 40)}...`);
        } else {
            console.log(`! 案例已存在: ${c.title.substring(0, 40)}... (Skip)`);
        }
    }

    console.log(`\\n✅ 完成！共添加 ${addedCount} 条案例`);
}

main()
    .catch((e) => {
        console.error(e);
        process.exit(1);
    })
    .finally(async () => {
        await prisma.$disconnect();
    });
'''
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(ts_content)
    
    print(f"已生成脚本: {output_file}")

def main():
    base_dir = '/Users/gaofeixiang/Desktop/公平竞争审查系统'
    
    # 处理文件1
    print("处理文件1: 全国公平竞争审查督查发现的问题政策措施.xlsx")
    file1_path = os.path.join(base_dir, '全国公平竞争审查督查发现的问题政策措施.xlsx')
    cases1 = process_file1(file1_path)
    print(f"  提取案例数: {len(cases1)}")
    
    generate_ts_script(
        cases1,
        '全国公平竞争审查督查发现的问题政策措施（浙江省）',
        '国家市场监督管理总局',
        '2024-12-01',
        os.path.join(base_dir, 'scripts', 'add_national_督查_cases.ts')
    )
    
    # 处理文件2
    print("\n处理文件2: 浙江省公平竞争审查第三方评估发现的妨碍统一市场和公平竞争问题.xlsx")
    file2_path = os.path.join(base_dir, '浙江省公平竞争审查第三方评估发现的妨碍统一市场和公平竞争问题.xlsx')
    cases2 = process_file2(file2_path)
    print(f"  提取案例数: {len(cases2)}")
    
    generate_ts_script(
        cases2,
        '浙江省公平竞争审查第三方评估发现的妨碍统一市场和公平竞争问题',
        '浙江省市场监督管理局',
        '2024-12-01',
        os.path.join(base_dir, 'scripts', 'add_zhejiang_评估_cases.ts')
    )
    
    print(f"\n总计生成 {len(cases1) + len(cases2)} 条案例的导入脚本")
    print("\n下一步请执行:")
    print("  npx ts-node scripts/add_national_督查_cases.ts")
    print("  npx ts-node scripts/add_zhejiang_评估_cases.ts")

if __name__ == '__main__':
    main()
